import time
import openpyxl
from tkinter import *
from tkinter import filedialog
from openpyxl.styles import Font
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

XLFILEPATH = ""
def openFile():
    global XLFILEPATH
    XLFILEPATH = filedialog.askopenfilename(initialdir=".", title="Select the excel file",filetypes= [("Excel File","*.xlsx")])
    return 0

window = Tk()
button = Button(text="Start",command=openFile)
button.pack()
window.mainloop()

""" Change the variables below to modify the script or I'll just input the variables from pytinker"""
TRNo = 9
TSCNo = 4


wb = openpyxl.load_workbook(XLFILEPATH)
sheet = wb['Sheet1']

browser = webdriver.Chrome()
browser.maximize_window()
browser.get("https://web.whatsapp.com")

search_xpath = "/html/body/div[1]/div/div/div[3]/div/div[1]/div/div/div[2]/div/div[2]"
searchbar = WebDriverWait(browser, 100).until(EC.presence_of_element_located((By.XPATH, search_xpath)))

titleRow = []
for col in range(TSCNo, sheet.max_column + 1):
    if (sheet.cell(TRNo, col).value) == None:
        continue
    titleRow.append(sheet.cell(TRNo, col).value)

sn1 = titleRow[0]
sn2 = titleRow[1]
sn3 = titleRow[2]
sn4 = titleRow[3]
sn5 = titleRow[4]
sn6 = titleRow[5]
sn7 = titleRow[6]


for row in range(11, sheet.max_row + 1):
    
    # This block reads the data from excel file and appends it into studentRow[] and makes sure to convert absent of A into 0
    SRSNo = 4
    studentRow = []
    studentRow.append(sheet.cell(row,2).value)
    studentRow.append(sheet.cell(row,3).value)
    for col in range(SRSNo, sheet.max_column + 1,2):
        if str(sheet.cell(row, col).value).isnumeric():
            studentRow.append(sheet.cell(row,col).value)
        else:
            studentRow.append(0)
           
    # Assigning the values to variables and can calulate averages here
    usn = studentRow[0]
    name = studentRow[1]
    sub1 = studentRow[2]
    sub2 = studentRow[3]
    sub3 = studentRow[4]
    sub4 = studentRow[5]
    sub5 = studentRow[6]
    sub6 = studentRow[7]
    sub7 = studentRow[8]
    attend = '0'
    avg = '0'
    phno = studentRow[9]

    print(studentRow)


    messege = format(str('Dear Parent, %0A Kindly Find CIE 1 Marks: %0A USN : {usn} %0A Name : {name} %0A {sn1} : {sub1}/30 %0A {sn2} : {sub2}/30 %0A {sn3} : {sub3}/40 %0A {sn4} : {sub4}/40 %0A {sn5} : {sub5}/40 %0A {sn6} : {sub6}/40 %0A {sn7} : {sub7}/40 %0A Regards, %0A Class Teacher %0A 3rd Sem AIML' )) 

    messege = messege.format(usn=usn,name=name,sub1=sub1,sub2=sub2,sub3=sub3,sub4=sub4,sub5=sub5,sub6=sub6,sub7=sub7,sn1=sn1, sn2=sn2, sn3=sn3, sn4=sn4, sn5=sn5, sn6=sn6, sn7=sn7)

    walink = "https://web.whatsapp.com/send?phone=+91"+str(phno)+"&text="+messege+"&app_absent=1"

    browser.get(walink)


    sendButton_xpath = '//*[@id="main"]/footer/div[1]/div/span[2]/div/div[2]/div[2]/button'
    try: 
            sendButton = WebDriverWait(browser, 15).until(EC.presence_of_element_located((By.XPATH, sendButton_xpath)))
            sendButton.click()
            break
    except:
            sheet.cell(row,20).value = "Error"
            sheet.cell(row,20).font = Font(color='00ff0000')
            with open("./error.txt","a") as file:
                    file.write(usn+"\n")
    
    time.sleep(3)

threedot_xpath = "/html/body/div[1]/div/div/div[3]/header/div[2]/div/span/div[4]/div/span"
logoutbutton_xpath = '/html/body/div[1]/div/div/div[3]/header/div[2]/div/span/div[4]/span/div/ul/li[6]/div'
confirm_xpath = "/html/body/div[1]/div/span[2]/div/div/div/div/div/div/div[3]/div/div[2]/div/div"

wb.save('studentData.xlsx')

browser.get("https://web.whatsapp.com/send?phone=+918050158461&app_absent=1")
threedot = WebDriverWait(browser, 15).until(EC.presence_of_element_located((By.XPATH, threedot_xpath)))
time.sleep(4)
threedot.click()
time.sleep(4)

logoutbutton = WebDriverWait(browser, 15).until(EC.presence_of_element_located((By.XPATH, logoutbutton_xpath)))
logoutbutton.click()

time.sleep(5)

confirmbutton = WebDriverWait(browser, 15).until(EC.presence_of_element_located((By.XPATH, confirm_xpath)))
time.sleep(4)
confirmbutton.click()

time.sleep(10)
browser.close()
browser.quit()
