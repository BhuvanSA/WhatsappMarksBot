from turtle import st
import openpyxl
import re


class ExcelManager:

    def __init__(self, filename: str, sheet_name: str):
        """Initializes the ExcelManager object"""
        self.sheet = openpyxl.load_workbook(filename)[sheet_name]
        self.__curr_row = 1
        self.__skip_header()
        self.column_names = []
        self.column_names_row = 0
        self.__get_column_names()
        self.max_marks = dict()
        self.max_internal = 1
        self.__get_max_marks()
        self.max_rows = self.sheet.max_row - self.__curr_row + 1
        # print(self.max_internal, self.max_marks)

    def __skip_header(self):
        ''' Skips the header by finding the row with 'Sl No' in the first column '''
        for row in range(1, 15):
            if str(self.sheet.cell(row, 1).value) == 'Sl No':
                self.__curr_row = row
                break

    def __get_column_names(self):
        ''' Gets the column names from the header row and stores it in self.column_names '''

        # Raise exception if the first column is not 'Sl No'
        if self.sheet.cell(self.__curr_row, 1).value != 'Sl No':
            raise Exception(
                'Sl No not found in the first column invalid __curr_row')
        else:
            self.column_names_row = self.__curr_row

        # Get the column names
        for col in range(1, self.sheet.max_column + 1):
            if (self.sheet.cell(self.__curr_row, col).value) == None:
                continue
            self.column_names.append(
                self.sheet.cell(self.__curr_row, col).value)

        # Increment the __curr_row
        self.__curr_row += 1

    def __get_max_marks(self):
        ''' Gets the max marks from the header row and stores it in self.max_marks '''
        for col in range(1, self.sheet.max_column + 1):
            value = str(self.sheet.cell(self.__curr_row, col).value)

            # Search for the pattern 'CIE1(100)' and store the value in self.max_marks
            if match := re.search(r'...(\d).*\((.+)\)', value):
                if match.group(1) == '1':
                    subject_code = str(self.sheet.cell(
                        self.__curr_row - 1, col).value)
                    self.max_marks[subject_code] = match.group(2)
                else:
                    self.max_internal = max(
                        self.max_internal, int(match.group(1)))

        # Increment the __curr_row
        self.__curr_row += 1

    def get_student_data(self, internal: int, slno: int = 1):
        """ Returns the student data in the form of a dictionary
            internal: int - The internal number (1, 2, 3, etc.)
            slno: int - The serial number of the student (1, 2, 3, etc.)
        """

        student_data = dict()
        row = self.__curr_row + slno - 1

        student_data['usn'] = str(self.sheet.cell(row, 2).value)
        if student_data['usn'] == 'None':
            return None
        student_data['name'] = str(self.sheet.cell(row, 3).value)

        for col in range(3 + internal, self.sheet.max_column - 1, self.max_internal):
            subject_code = str(self.sheet.cell(
                self.column_names_row, col).value)
            value = str(self.sheet.cell(row, col).value)
            student_data[subject_code] = f"{value}/{self.max_marks[subject_code]}"

        student_data['phone_number'] = str(
            self.sheet.cell(row, self.sheet.max_column).value)

        return student_data
