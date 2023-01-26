
import openpyxl
wb = openpyxl.load_workbook('studentData.xlsx')
sheet = wb['Sheet1']

sn1 = sheet.cell(1,3).value
sn2 = sheet.cell(1,4).value
sn3 = sheet.cell(1,5).value
sn4 = sheet.cell(1,6).value
sn5 = sheet.cell(1,7).value
sn6 = sheet.cell(1,8).value
sn7 = sheet.cell(1,9).value
sn8 = sheet.cell(1,10).value


for row in range(2, sheet.max_row + 1):
    usn = sheet.cell(row,1).value
    name = sheet.cell(row,2).value
    sub1 = sheet.cell(row,3).value
    sub2 = sheet.cell(row,4).value
    sub3 = sheet.cell(row,5).value
    sub4 = sheet.cell(row,6).value
    sub5 = sheet.cell(row,7).value
    sub6 = sheet.cell(row,8).value
    sub7 = sheet.cell(row,9).value
    sub8 = sheet.cell(row,10).value
    attend = sheet.cell(row,11).value
    phno = sheet.cell(row,12).value
    avg = (sub1+sub2+sub3+sub4+sub5+sub6+sub7+sub8)/8


    # messege = str(usn+" "+name+" "+str(sub1)+" "+str(sub2)+" "+str(sub3)+" "+str(sub4)+" "+str(sub5)+" "+str(sub6)+" "+str(sub7)+" "+str(sub8)+" "+str(avg)+" "+str(attend))


    messege = format(str('Dear Parent,\nUSN : {usn}\nName : {name}\n{sn1} : {sub1}\n{sn2} : {sub2}\n{sn3} : {sub3}\n{sn4} : {sub4}\n{sn5} : {sub5}\n{sn6} : {sub6}\n{sn7} : {sub7}\n{sn8} : {sub8}\nAverage Marks: {avg}\nAttendance : {attend}' )) 

    messege = messege.format(usn=usn,name=name,sub1=sub1,sub2=sub2,sub3=sub3,sub4=sub4,sub5=sub5,sub6=sub6,sub7=sub7,sub8=sub8,avg=avg,attend=attend, sn1=sn1, sn2=sn2, sn3=sn3, sn4=sn4, sn5=sn5, sn6=sn6, sn7=sn7, sn8=sn8)

    print(messege)

