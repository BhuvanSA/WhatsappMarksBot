import openpyxl
import pywhatkit

# Read student marks, attendance and parent phone numbers from Excel spreadsheet
wb = openpyxl.load_workbook('student_marks.xlsx')
sheet = wb['Sheet1']

# Iterate through rows in the spreadsheet
for row in range(2, sheet.max_row + 1):
    student_id = sheet.cell(row, 1).value
    student_name = sheet.cell(row, 2).value
    student_marks = []
    for col in range(3, 11):
        student_marks.append(sheet.cell(row, col).value)
    student_attendance = sheet.cell(row, 11).value
    student_parent_number = sheet.cell(row, 12).value
    # Create a table of marks
    marks_table = f'Subject\t\tMarks\n'
    marks_table += f'Subject 1\t\t{student_marks[0]}\n'
    marks_table += f'Subject 2\t\t{student_marks[1]}\n'
    marks_table += f'Subject 3\t\t{student_marks[2]}\n'
    marks_table += f'Subject 4\t\t{student_marks[3]}\n'
    marks_table += f'Subject 5\t\t{student_marks[4]}\n'
    marks_table += f'Subject 6\t\t{student_marks[5]}\n'
    marks_table += f'Subject 7\t\t{student_marks[6]}\n'
    marks_table += f'Subject 8\t\t{student_marks[7]}\n'
    # Send report of marks and attendance to student's parent via WhatsApp
    pywhatkit.sendwhatmsg("+14155238886", student_parent_number, "Report of marks for Student ID: "+student_id+" - "+student_name+":\n"+marks_table+"\nAttendance: "+str(student_attendance), 0, 0)
    print(f'Sent message to {student_parent_number}')
