import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

wb = openpyxl.load_workbook('studentData.xlsx')
sheet = wb['Sheet1']

browser = webdriver.Chrome(
        executable_path="/Users/bhuvansa/chromedriver/chromedriver")
browser.maximize_window()
browser.get("https://web.whatsapp.com")

search_xpath = "/html/body/div[1]/div/div/div[3]/div/div[1]/div/div/div[2]/div/div[2]"
searchbar = WebDriverWait(browser, 100).until(
        EC.presence_of_element_located((By.XPATH, search_xpath))
    )
sn1 = sheet.cell(1,3).value
sn2 = sheet.cell(1,4).value
sn3 = sheet.cell(1,5).value
sn4 = sheet.cell(1,6).value
sn5 = sheet.cell(1,7).value
sn6 = sheet.cell(1,8).value
sn7 = sheet.cell(1,9).value
sn8 = sheet.cell(1,10).value


for row in range(2, sheet.max_row + 1):
        
    usn = sheet.cell(row,1).value
    name = sheet.cell(row,2).value
    sub1 = sheet.cell(row,3).value
    sub2 = sheet.cell(row,4).value
    sub3 = sheet.cell(row,5).value
    sub4 = sheet.cell(row,6).value
    sub5 = sheet.cell(row,7).value
    sub6 = sheet.cell(row,8).value
    sub7 = sheet.cell(row,9).value
    sub8 = sheet.cell(row,10).value
    attend = sheet.cell(row,11).value
    phno = sheet.cell(row,12).value
    avg = (sub1+sub2+sub3+sub4+sub5+sub6+sub7+sub8)/8


    # messege = str(usn+" "+name+" "+str(sub1)+" "+str(sub2)+" "+str(sub3)+" "+str(sub4)+" "+str(sub5)+" "+str(sub6)+" "+str(sub7)+" "+str(sub8)+" "+str(avg)+" "+str(attend))


    messege = format(str('Dear Parent,\nUSN : {usn}\nName : {name}\n{sn1} : {sub1}\n{sn2} : {sub2}\n{sn3} : {sub3}\n{sn4} : {sub4}\n{sn5} : {sub5}\n{sn6} : {sub6}\n{sn7} : {sub7}\n{sn8} : {sub8}\nAverage Marks: {avg}\nAttendance : {attend}' )) 

    messege = messege.format(usn=usn,name=name,sub1=sub1,sub2=sub2,sub3=sub3,sub4=sub4,sub5=sub5,sub6=sub6,sub7=sub7,sub8=sub8,avg=avg,attend=attend, sn1=sn1, sn2=sn2, sn3=sn3, sn4=sn4, sn5=sn5, sn6=sn6, sn7=sn7, sn8=sn8)




    walink = "https://web.whatsapp.com/send?phone="+str(phno)+"&text="+messege+"&app_absent=1"

    browser.get(walink)

    #Dynamic Code
    sendButton_xpath = '//*[@id="main"]/footer/div[1]/div/span[2]/div/div[2]/div[2]/button'
    
    sendButton = WebDriverWait(browser, 150).until(
            EC.presence_of_element_located((By.XPATH, sendButton_xpath))
        )
    
    sendButton.click()
    time.sleep(2)

threedot_xpath = "/html/body/div[1]/div/div/div[3]/header/div[2]/div/span/div[4]/div/span"
logoutbutton_xpath = '/html/body/div[1]/div/div/div[3]/header/div[2]/div/span/div[4]/span/div/ul/li[6]/div'
confirm_xpath = "/html/body/div[1]/div/span[2]/div/div/div/div/div/div/div[3]/div/div[2]/div/div"

threedot = WebDriverWait(browser, 5).until(
        EC.presence_of_element_located((By.XPATH, threedot_xpath))
    )
threedot.click()
time.sleep(3)

logoutbutton = WebDriverWait(browser, 5).until(
        EC.presence_of_element_located((By.XPATH, logoutbutton_xpath))
    )
logoutbutton.click()

time.sleep(3)

confirmbutton = WebDriverWait(browser, 5).until(
        EC.presence_of_element_located((By.XPATH, confirm_xpath))
    )
confirmbutton.click()

time.sleep(10)
browser.close()
browser.quit()

