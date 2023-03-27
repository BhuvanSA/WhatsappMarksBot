# Importing all required libararies

# time to add waittime
import time
import os
import sys

# To read data from excel
import openpyxl
from openpyxl.styles import Font  # To style text

# tkinter for window based management
from tkinter import *
from tkinter import filedialog

# selenium for browser stuff
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# PIL
from PIL import ImageTk, Image

# Global variable

# Default Values to Start Reading Data

# A function which checks weather the given word contains 30 or 40
# If so return it else return 0


def regx30or40(word):
    if word != None:
        for i in word:
            if i == '3':
                return 30
            elif i == '4':
                return 40

    return 0


# The core of the program which does most of the work by calling other functions too
def brain():

    XLFILEPATH = filedialog.askopenfilename(
        initialdir=".", title="Select the excel file", filetypes=[("Excel File", "*.xlsx")])
    wb = openpyxl.load_workbook(XLFILEPATH)
    print('loaded')
    sheet = wb['Sheet1']
    subjectRow = []
    print('loaded sheet')

    row = 0
    # Read all header names in Sl No col and store it in subjectRow
    for row in range(1, sheet.max_row + 1):
        if str(sheet.cell(row, 1).value) == 'Sl No':
            for col in range(2, sheet.max_column + 1):
                if (sheet.cell(row, col).value) == None:
                    continue
                subjectRow.append(sheet.cell(row, col).value)
            break
    row += 1
    submarksRow = []
    print('loaded sheet and read all header names')


# This needs to be changed to work for 3nd IA
    # use regex to find the marks inside the brackets and append it into a new array
    for col in range(4, sheet.max_column-1, 3):
        if regx30or40(sheet.cell(row, col).value) != 0:
            submarksRow.append(regx30or40(sheet.cell(row, col).value))

    # Browser init
    browser = webdriver.Chrome()
    browser.maximize_window()

    # Head to whatsapp.com to allow user to login
    browser.get("https://web.whatsapp.com")

    # Wait for user to login to their account
    search_xpath = "/html/body/div[1]/div/div/div[4]/div/div[1]/div/div/div[2]/div/div[1]/p"
    searchbar = WebDriverWait(browser, 200).until(
        EC.presence_of_element_located((By.XPATH, search_xpath)))

    # Read all student data and store it in a list
    for row2 in range(row+1, sheet.max_row + 1):
        studentRow = []
        if str(sheet.cell(row2, 1).value).isnumeric():
            studentRow.append(sheet.cell(row2, 2).value)  # Storing USN
            studentRow.append(sheet.cell(row2, 3).value)  # Stroing StudentName

# This needs to be changed to work for 3nd IA
            # Storing all subjects marks data starting from 5 skipping 2
            for col in range(5, sheet.max_column + 1, 3):
                if str(sheet.cell(row2, col).value).isnumeric():
                    studentRow.append(sheet.cell(row2, col).value)
                else:
                    studentRow.append(0)

            # Generate a messege base on the data
            messege = messegeGenerator(subjectRow, studentRow, submarksRow)
            walink = "https://web.whatsapp.com/send?phone=+91" + \
                str(studentRow[-1])+"&text="+messege+"&app_absent=1"

            # Access the parents no and send the messege
            browser.get(walink)

            sendButton_xpath = '//*[@id="main"]/footer/div[1]/div/span[2]/div/div[2]/div[2]/button'
            try:
                sendButton = WebDriverWait(browser, 25).until(
                    EC.presence_of_element_located((By.XPATH, sendButton_xpath)))
                sendButton.click()
            except:
                sheet.cell(row, sheet.max_column+1).value = "Error"
                sheet.cell(row, sheet.max_column +
                           1).font = Font(color='00ff0000')

            time.sleep(5)

    # X-paths for three buttons
    threedot_xpath = "/html/body/div[1]/div/div/div[3]/header/div[2]/div/span/div[4]/div/span"
    logoutbutton_xpath = '/html/body/div[1]/div/div/div[3]/header/div[2]/div/span/div[4]/span/div/ul/li[6]/div'
    confirm_xpath = "/html/body/div[1]/div/span[2]/div/div/div/div/div/div/div[3]/div/div[2]/div/div"

    # Saving all the changes made to studentData
    wb.save('studentData.xlsx')

    # Head to my whatsapp no and logout
    browser.get("https://web.whatsapp.com/send?phone=+918050158461&app_absent=1")
    threedot = WebDriverWait(browser, 15).until(
        EC.presence_of_element_located((By.XPATH, threedot_xpath)))
    time.sleep(4)
    threedot.click()
    time.sleep(4)

    logoutbutton = WebDriverWait(browser, 15).until(
        EC.presence_of_element_located((By.XPATH, logoutbutton_xpath)))
    logoutbutton.click()

    time.sleep(5)

    confirmbutton = WebDriverWait(browser, 15).until(
        EC.presence_of_element_located((By.XPATH, confirm_xpath)))
    time.sleep(4)
    confirmbutton.click()

    time.sleep(10)
    browser.close()
    browser.quit()

# A function which accepts three lists and return a messege to send to parents


def messegeGenerator(nameList, detailList, SML):
    messege = format(str(
        'Dear Parent, %0A Kindly Find CIE 2 Marks: %0A {unn} : {usn} %0A {nme} : {name} %0A {sn1} : {sub1}/{sm1} %0A {sn2} : {sub2}/{sm2} %0A {sn3} : {sub3}/{sm3} %0A {sn4} : {sub4}/{sm4} %0A {sn5} : {sub5}/{sm5} %0A {sn6} : {sub6}/{sm6} %0A {sn7} : {sub7}/{sm7} %0A Regards, %0A Class Teacher %0A 3rd Sem AIML'))
    messege = messege.format(sm1=SML[0], sm2=SML[1], sm3=SML[2], sm4=SML[3], sm5=SML[4], sm6=SML[5], sm7=SML[6], unn=nameList[0], usn=detailList[0], nme=nameList[1], name=detailList[1], sub1=detailList[2], sub2=detailList[3],
                             sub3=detailList[4], sub4=detailList[5], sub5=detailList[6], sub6=detailList[7], sub7=detailList[8], sn1=nameList[2], sn2=nameList[3], sn3=nameList[4], sn4=nameList[5], sn5=nameList[6], sn6=nameList[7], sn7=nameList[8])
    return messege

# for the sake of it


def main():
    # try:
    brain()
    # except:
    # print("Something went wrong contact BhuvanSA")

    print("main final done")


root = Tk(className="Whatsapp Marks Bot")
root.geometry("600x400")
root.configure(bg="lightgrey")

imgFrame = Frame(root, width=600, height=99)
imgFrame.grid(row=0)

bundle_dir = getattr(
    sys, 'head.jpg', os.path.abspath(os.path.dirname(__file__)))
# path_to_help = os.path.abspath(os.path.join(bundle_dir, 'help.md'))

img = ImageTk.PhotoImage(Image.open("Resources/head.jpg"))

imgLabel = Label(imgFrame, image=img, pady=0, padx=0)
imgLabel.grid(row=0)

headlabel = Label(root, padx=5, text="Welcome to Whatsapp Marks Bot", font=(
    "", 30), fg="black", bg="lightgrey")
headlabel.grid(row=1)


btnLabel = Label(root, text="Select the Excel file to start",
                 font=("", 30), bg="lightgrey", fg="black")
btnLabel.grid(row=2)


button = Button(root, padx=5, text="Select", command=main,
                font=("", 24), bg="white", fg="black")
button.grid(row=3)

root.mainloop()
