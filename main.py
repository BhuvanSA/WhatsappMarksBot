import time
import openpyxl
from openpyxl.styles import Font
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

wb = openpyxl.load_workbook('studentData.xlsx')
sheet = wb['Sheet1']

browser = webdriver.Chrome()
browser.maximize_window()
browser.get("https://web.whatsapp.com")

search_xpath = "/html/body/div[1]/div/div/div[3]/div/div[1]/div/div/div[2]/div/div[2]"
searchbar = WebDriverWait(browser, 100).until(
        EC.presence_of_element_located((By.XPATH, search_xpath))
    )
sn1 = sheet.cell(9,4).value
sn2 = sheet.cell(9,6).value
sn3 = sheet.cell(9,8).value
sn4 = sheet.cell(9,10).value
sn5 = sheet.cell(9,12).value
sn6 = sheet.cell(9,14).value
sn7 = sheet.cell(9,16).value


for row in range(11, sheet.max_row + 1):

    if row > 13:
        continue

    usn = sheet.cell(row,2).value
    name = sheet.cell(row,3).value
    sub1 = sheet.cell(row,4).value
    sub2 = sheet.cell(row,6).value
    sub3 = sheet.cell(row,8).value
    sub4 = sheet.cell(row,10).value
    sub5 = sheet.cell(row,12).value
    sub6 = sheet.cell(row,14).value
    sub7 = sheet.cell(row,16).value
    attend = '0'
    avg = '0'
    phno = sheet.cell(row,19).value


    messege = format(str('Dear Parent, %0A Kindly Find CIE 1 Marks: %0A USN : {usn} %0A Name : {name} %0A {sn1} : {sub1}/30 %0A {sn2} : {sub2}/30 %0A {sn3} : {sub3}/40 %0A {sn4} : {sub4}/40 %0A {sn5} : {sub5}/40 %0A {sn6} : {sub6}/40 %0A {sn7} : {sub7}/40 %0A Regards, %0A Class Teacher %0A 3rd Sem AIML' )) 

    messege = messege.format(usn=usn,name=name,sub1=sub1,sub2=sub2,sub3=sub3,sub4=sub4,sub5=sub5,sub6=sub6,sub7=sub7,sn1=sn1, sn2=sn2, sn3=sn3, sn4=sn4, sn5=sn5, sn6=sn6, sn7=sn7)

    walink = "https://web.whatsapp.com/send?phone=+91"+str(phno)+"&text="+messege+"&app_absent=1"

    browser.get(walink)


    sendButton_xpath = '//*[@id="main"]/footer/div[1]/div/span[2]/div/div[2]/div[2]/button'
    try: 
            sendButton = WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.XPATH, sendButton_xpath)))
            # sendButton.click()
    except:
            sheet.cell(row,20).value = "Error"
            sheet.cell(row,20).font = Font(color='00ff0000')
            with open("./error.txt","a") as file:
                    file.write(usn+"\n")
    
    time.sleep(3)

threedot_xpath = "/html/body/div[1]/div/div/div[3]/header/div[2]/div/span/div[4]/div/span"
logoutbutton_xpath = '/html/body/div[1]/div/div/div[3]/header/div[2]/div/span/div[4]/span/div/ul/li[6]/div'
confirm_xpath = "/html/body/div[1]/div/span[2]/div/div/div/div/div/div/div[3]/div/div[2]/div/div"

wb.save('studentData.xlsx')

threedot = WebDriverWait(browser, 5).until(
        EC.presence_of_element_located((By.XPATH, threedot_xpath))
    )
threedot.click()
time.sleep(3)

logoutbutton = WebDriverWait(browser, 5).until(
        EC.presence_of_element_located((By.XPATH, logoutbutton_xpath))
    )
logoutbutton.click()

time.sleep(3)

confirmbutton = WebDriverWait(browser, 5).until(
        EC.presence_of_element_located((By.XPATH, confirm_xpath))
    )
confirmbutton.click()

time.sleep(10)
browser.close()
browser.quit()
